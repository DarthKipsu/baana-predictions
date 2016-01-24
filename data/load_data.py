from openpyxl import load_workbook

def read_xlsx_file(path, sheet):
    return load_workbook(path).get_sheet_by_name(sheet)

def print_content(path, sheet_name, cells):
    sheet = read_xlsx_file(path, sheet_name)
    for row in sheet.iter_rows(cells):
        print(row[0].value, row[1].value, row[2].value)

# currently the data is in rows B3:D850
print_content('original/Baanan_pyorailijamaarat.xlsx', 'Taul1', 'B3:D50')
