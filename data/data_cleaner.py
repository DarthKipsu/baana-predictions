from bs4 import BeautifulSoup
from openpyxl import load_workbook

import numpy as np

def read_xlsx_file(path, sheet):
    """
    Takes a path to xlsx (Excel) file and a sheet name within the file and returns an openpyxl
    Workbook sheet.
    """
    return load_workbook(path).get_sheet_by_name(sheet)

def read_xml_file(path):
    """
    Takes a path to xml file and returns BeautifulSoup object representing the document.
    """
    return BeautifulSoup(open(path), 'lxml-xml')

def write_to_file(path, data):
    """
    Takes a path to file and data to br written as 2d array.
    Overwrites the file with the given data.
    """
    f = open(path, 'w+')
    for row in data:
        for column in row:
            f.write("%s " % column)
        f.write('\n')
    f.close()

def daily_counts_from_hourly(raw_data, count_func):
    """
    Takes original data with hourly cycler counts and compresses them to daily counts. Uses
    count_func to read the correct cyclist value from the data.
    """
    data = []
    hour = 0
    count = 0
    for row in raw_data:
        count += count_func(row)
        hour += 1
        if (hour == 24):
            data.append([count])
            hour = 0
            count = 0
    return data

def clean_labels():
    """
    Reads labels from cyclics count files and writes them on a file without all the unneeded extra
    information.
    """
    # Read data from 1.1.2013 until 28.4.2015
    sheet = read_xlsx_file('original/Baanan_pyorailijamaarat.xlsx', 'Taul1')
    data = [[row[2].value] for row in sheet.iter_rows('B3:D850')]

    # Read data between 29.4.2015 and 30.9.2015
    with open('original/Helsingin_pyorailijamaarat.csv') as csv_file:
        csv_data = np.array([row.split(',') for row in csv_file.readlines()])
        clean_data = daily_counts_from_hourly(csv_data[11593:15313], lambda x: int(x[16]))
        data.extend(clean_data)

    # Read data between 1.10.2015 and 31.12.2015
    sheet = read_xlsx_file('original/Baana_syys_joulukuu.xlsx', 'Pohjoinen Rautatiekatu')
    clean_data = daily_counts_from_hourly(sheet.iter_rows('A2:B2209'), lambda x: x[1].value)
    data.extend(clean_data)
    return data

def clean_weather_data(path):
    """
    Reads data from given xml file and returns it as a clean array of floats. The array contains a
    row for each day with the columns containing:
    col 0: amount of rain
    col 1: medium temperature
    col 2: snow depth
    col 3: minimum temperature
    col 4: maximum temperature
    """
    soup = read_xml_file(path)
    data = []
    i = 0
    for feature in soup.find_all('MeasurementTimeseries'):
        j = 0
        feature_name = feature.get('gml:id')
        for value_tag in feature.find_all('value'):
            value = float(value_tag.string)
            if value != value:
                value = -1.0
            if i == 0:
                data.append([value])
            else:
                data[j].append(value)
            j += 1
        i += 1
    return data

def clean_original_data_files():
    data = clean_weather_data('original/fmi-2013-2014.xml')
    data.extend(clean_weather_data('original/fmi-2014-2015.xml'))
    data.extend(clean_weather_data('original/fmi-2015-2016.xml'))
    data.extend(clean_weather_data('original/fmi-2016-beginning.xml'))
    
    write_to_file('clean/data', data)
    write_to_file('clean/labels', clean_labels())
